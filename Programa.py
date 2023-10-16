import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import warnings
warnings.filterwarnings("ignore")


def get_excel_files() -> list:
    """
    Obtiene una lista de los nombres de archivos Excel (.xlsx) en la carpeta donde se encuentran la nueva data descargada.

    Returns:
    excel_files: Una lista de nombres de archivos Excel con extensión .xlsx encontrados en la carpeta.
    """
     # Lista todos los archivos en la ubicación dada
    all_files = os.listdir(os.path.join("C:\\Users","Marcos","Desktop", "Meses"))

    # Inicializa una lista vacía para almacenar los nombres de archivos Excel
    excel_files = []

    # Itera sobre todos los archivos encontrados
    for file in all_files:
        # Verifica si el archivo tiene una extensión .xlsx
        if file.endswith(".xlsx"):
            # Agrega la ruta completa del archivo a la lista excel_files
            excel_files.append("C:/Users/Marcos/Desktop/Meses/" + file)

    return excel_files


def nota_de_credito(data_nueva: pd.DataFrame) -> pd.DataFrame:
    """
    Procesa un DataFrame de ventas y ajusta los valores de 'Imp. Total' para las 'Notas de Crédito'.
    
    Args:
    data_nueva (pd.DataFrame): El DataFrame que contiene todos los datos nuevos.

    Returns:
    pd.DataFrame: El DataFrame modificado con los ajustes de las 'Notas de Crédito'.
    """
    # Itera a través de las filas del DataFrame
    for i in range(len(data_nueva)):
        # Verifica si la columna en la posición 1 contiene la cadena "Nota de Crédito"
        if "Nota de Crédito" in data_nueva.iloc[i,1]:
            # Ajusta el valor en la columna 4 multiplicándolo por -1
            data_nueva.iloc[i,4] = data_nueva.iloc[i,4]*-1
    
    return data_nueva


def chequeo_duplicados(data_Excel_completa, data_nueva):
    """
    Comprueba y elimina aquellas facturas que ya se encuentran en el excel para evitar duplicados.

    Args:
    data_Excel_completa: El DataFrame con todas las facturas encontradas en el Excel.
    data_nueva (pd.DataFrame): El DataFrame con todas las facturas nuevas que se quieren agregar al Excel.

    Returns:
    pd.DataFrame: Un DataFrame con todas facturas que no se encuentran en el Excel.
    """
    # Selecciona la columna que indica el número de factura de ambos DataFrames
    column_data_Excel_completa = data_Excel_completa.iloc[:,2]
    column_data_nueva = data_nueva.iloc[:,2]
    # Encuentra los índices de filas duplicadas en 'data_nueva' en comparación con 'data_Excel_completa'
    indices_duplicados = column_data_nueva[column_data_nueva.isin(column_data_Excel_completa)].index

    # Elimina las filas duplicadas de 'data_nueva' basadas en los índices encontrados
    data_sin_duplicados = (data_nueva.drop(index=indices_duplicados)).reset_index(drop=True)

    return data_sin_duplicados


def main_menu():
    """
    Le muestra al usuario las acciones que el progrmaa puede realizar y le pide que seleccione que desea hacer.
    """
    while True:
        os.system('cls') # Limpia la pantalla de la consola
        print("Menu Principal:\n1) Actualizar Excel de Cliente\n2) Crear Excel Para Nuevo Cliente\n3) Salir del programa")
        choice = input("Ingrese un Número: ") # Solicita la elección del usuario

        if choice == "1":
            os.system('cls')
            print("Has Seleccionado la Opcion 'Actualizar Excel de Cliente'. Usted Va a Actualizar los Excels de los Siguientes Clientes:")
            
            # Lee la lista de clientes y archivos
            cuits = pd.read_excel('C:/Users/Marcos/Desktop/PRUEBA CUITS.xlsx')
            file_paths = get_excel_files()
            clientes_cuits = []

            # Encuentra clientes relacionados con los archivos existentes
            for c in range(len(cuits)):
                for file_path in file_paths:
                    if str(cuits.iloc[c,1]) in file_path:
                        clientes_cuits.append(cuits.iloc[c,0])
            clientes_cuits = list(set(clientes_cuits)) # Elimina duplicados

            for cliente in clientes_cuits:
                print(cliente)

            print("\nDesea continuar?\n1)Si, actualizarlos\n2)No, volver al menú")
            eleccion=input("Ingrese un Número: ")

            if eleccion == "1":
                actualizacion_excel()
                os.system('cls')
                print("Los Excels de los siguientes clientes han sido actualizados:")
                
                for cliente in clientes_cuits:
                    print(cliente)
                
                asd = input("\nVolver al menú?\n1)Si\n2)No, salir del programa\nIngrese un Número: ")
                
                if asd !="1":
                    break
            else:
                os.system('cls')
        
        elif choice == "2":
            os.system('cls')
            print("Has Seleccionado la Opcion 'Crear Excel Para Nuevo Cliente'.")
            nuevo_excel()
            os.system('cls')
            
            abc = input("El excel del cliente ha sido creado. Volver al menú?\n1)Si\n2)No, salir del programa\nIngrese un Número: ")
            
            if abc != "1":
                break
        else:
            break


def actualizacion_excel():
    """
    Actualiza todos los excels de todos los clientes disponibles para actualizar, es decir, de aquellos clientes que se 
    tenga nueva información para agregar a sus respectivos excels.
    """
    cuits = pd.read_excel('C:/Users/Marcos/Desktop/PRUEBA CUITS.xlsx') # Lee la lista de clientes
    file_paths = get_excel_files() # Obtiene la lista de rutas de archivos Excel
    
    i=0 # Contador de archivos procesados

    for c in range(len(cuits)): # Itera sobre la lista de clientes
        a=0

        # Comprueba si aún quedan archivos por procesar
        if i != len(file_paths):
            # Inicializa DataFrames para almacenar datos de facturas emitidas y recibidas
            alldata_emitidos = pd.DataFrame()
            alldata_recibidos = pd.DataFrame()

            # Itera sobre los archivos buscando coincidencias con el cliente actual
            for file_path in file_paths:
                # Comprueba si el CUIT del cliente está en la ruta del archivo
                if str(cuits.iloc[c,1]) in file_path:
                    if 'Emitidos' in file_path:
                        # Lee y concatena datos de facturas emitidas
                        alldata_emitidos = pd.concat([alldata_emitidos, pd.read_excel(file_path,skiprows=1)], axis=0, ignore_index=True).reset_index(drop=True)
                    else:
                        # Lee y concatena datos de facturas recibidas
                        alldata_recibidos = pd.concat([alldata_recibidos, pd.read_excel(file_path,skiprows=1)], axis=0, ignore_index=True).reset_index(drop=True)  
                    i+=1
            
            if len(alldata_emitidos) > 0:
                # Filtra y selecciona los datos de ventas del DataFrame completo.
                data_nueva = alldata_emitidos[['Fecha', 'Tipo', 'Número Desde', 'Denominación Receptor', 'Imp. Total']]

                # Convierte la columna 'Fecha' a formato de fecha.
                data_nueva['Fecha'] = pd.to_datetime(data_nueva['Fecha'], format='%d/%m/%Y')
                data_nueva['Fecha'] = data_nueva['Fecha'].dt.strftime('%d/%m/%Y') 

                # Ordena el dataframe por número de factura
                data_nueva = data_nueva.sort_values(by='Número Desde', ascending=True).reset_index(drop=True)

                # Aplica la función 'nota_de_credito' a los datos de ventas.
                data_nueva = nota_de_credito(data_nueva)

                # Lee los datos existentes del archivo Excel del cliente.
                excel_filepath = f'C:/Users/Marcos/Desktop/Oficina/Monotributo/{cuits.iloc[c, 0]}.xlsx'
                data_excel = pd.read_excel(excel_filepath, sheet_name="VENTAS NUEVO", usecols="A:E", skiprows=5)

                # Aplica la función 'chequeo_duplicados' a los datos de ventas.
                data_nueva = chequeo_duplicados(data_excel, data_nueva)

                # Calcula el tamaño actual de los datos en el archivo Excel.
                size = len(data_excel)

                # Abre el archivo Excel y selecciona la hoja de trabajo 'VENTAS NUEVO'.
                wb = openpyxl.load_workbook(excel_filepath)
                sheet = wb['VENTAS NUEVO']

                # Agrega datos actualizados al archivo Excel
                for row_idx, row in enumerate(openpyxl.utils.dataframe.dataframe_to_rows(data_nueva, index=False, header=False), (size+7)):
                    for col_idx, value in enumerate(row, 1):
                        sheet.cell(row=row_idx, column=col_idx, value=value)

                # Aplica formato a las celdas para mejorar la apariencia.
                for row in wb["VENTAS NUEVO"].iter_rows(min_row=(size+7), max_row=(size+len(data_nueva)+7), min_col=1, max_col=5):
                    for cell in row:
                        cell.font = Font(name="Calibri", size=11) 
                        if cell.column == 1:
                            cell.alignment = Alignment(horizontal="right")  
                        elif cell.column == 3:
                            cell.alignment = Alignment(horizontal="center")  
                wb.save(excel_filepath)

            if len(alldata_recibidos) > 0:
                # Filtra y selecciona los datos de compras del DataFrame completo.
                data_nueva = alldata_recibidos[['Fecha', 'Tipo', 'Número Desde', 'Denominación Emisor', 'Imp. Total']]

                # Convierte la columna 'Fecha' a formato de fecha y ordena el dataframe.
                data_nueva['Fecha'] = pd.to_datetime(data_nueva['Fecha'], format='%d/%m/%Y')
                data_nueva = data_nueva.sort_values(by='Fecha', ascending=True).reset_index(drop=True)
                data_nueva['Fecha'] = data_nueva['Fecha'].dt.strftime('%d/%m/%Y') 

                # Aplica la función 'nota_de_credito' a los datos de compras.
                data_nueva = nota_de_credito(data_nueva)

                # Lee los datos existentes del archivo Excel del cliente.
                excel_filepath = f'C:/Users/Marcos/Desktop/Oficina/Monotributo/{cuits.iloc[c, 0]}.xlsx'
                data_excel = pd.read_excel(excel_filepath, sheet_name="COMPRAS NUEVO", usecols="A:E", skiprows=5)

                # Aplica la función 'chequeo_duplicados' a los datos de compras.
                data_nueva = chequeo_duplicados(data_excel, data_nueva)

                # Calcula el tamaño actual de los datos en el archivo Excel.
                size = len(data_excel)

                # Abre el archivo Excel y selecciona la hoja de trabajo 'COMPRAS NUEVO'.
                wb = openpyxl.load_workbook(excel_filepath)
                sheet = wb['COMPRAS NUEVO']

                # Agrega datos actualizados al archivo Excel
                for row_idx, row in enumerate(openpyxl.utils.dataframe.dataframe_to_rows(data_nueva, index=False, header=False), (size+7)):
                    for col_idx, value in enumerate(row, 1):
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Formatea la apariencia de las celdas
                for row in wb["COMPRAS NUEVO"].iter_rows(min_row=(size+7), max_row=(size+len(data_nueva)+7), min_col=1, max_col=5):
                    for cell in row:
                        cell.font = Font(name="Calibri", size=11) 
                        if cell.column == 1:
                            cell.alignment = Alignment(horizontal="right")  
                wb.save(excel_filepath)
        else:
            break


def nuevo_excel():
    """
    Crea un nuevo excel para un nuevo cliente.
    """
    nene=0
    while True:
        cuits = pd.read_excel('C:/Users/Marcos/Desktop/PRUEBA CUITS.xlsx') # Lee la lista de clientes desde un archivo Excel
        cuits2 = cuits # Realiza una copia de la lista de clientes

        # Solicita al usuario el nombre y el CUIT del nuevo cliente
        nuevo_cliente = pd.Series([input("Ingrese el Nombre del Nuevo Cliente: ").upper(),
                                input("Ingrese el CUIT del Nuevo Cliente: ")],
                                index=cuits.columns)

        # Agrega el nuevo cliente a la lista de clientes
        cuits2 = cuits2.append(nuevo_cliente, ignore_index=True)

        # Pregunta al usuario si desea crear el Excel o hacer otros cambios
        eleccion = input(f"\nUsted está por crear un Excel para el cliente {cuits2.iloc[len(cuits2)-1,0]} y su CUIT es {cuits2.iloc[len(cuits2)-1,1]}\nQue desea hacer?\n1)Crear Excel\n2)Cambiar el Nombre o el CUIT\n3)Volver al menú\nIngrese un Número:")
        
        if eleccion == "1":
             # Si el usuario elige crear un Excel, actualiza la lista de clientes
            cuits = cuits2
            cuits.to_excel('C:/Users/Marcos/Desktop/PRUEBA CUITS.xlsx', index=False)
            # Sale del bucle
            break
        elif eleccion == "3":
            # Si el usuario elige volver al menú, establece una bandera (nene) para salir del bucle principal
            nene=1
            break
        else:
            os.system('cls') # Limpia la pantalla si el usuario quiere volver a ingresar los datos del nuevo cliente

    if nene == 0:
        file_paths = get_excel_files() # Obtiene la lista de rutas de archivos Excel

        # Inicializa DataFrames para almacenar datos de facturas emitidas y recibidas
        alldata_emitidos = pd.DataFrame()
        alldata_recibidos = pd.DataFrame()

        # Itera sobre los archivos buscando coincidencias con el cliente actual
        for file_path in file_paths:
            # Comprueba si el CUIT del cliente está en la ruta del archivo
            if str(cuits.iloc[(len(cuits)-1),1]) in file_path:
                if 'Emitidos' in file_path:
                    # Lee y concatena datos de facturas emitidas
                    alldata_emitidos = pd.concat([alldata_emitidos, pd.read_excel(file_path,skiprows=1)], axis=0, ignore_index=True).reset_index(drop=True)
                else:
                    # Lee y concatena datos de facturas recibidas
                    alldata_recibidos = pd.concat([alldata_recibidos, pd.read_excel(file_path,skiprows=1)], axis=0, ignore_index=True).reset_index(drop=True)  
        
        source_file = "C:/Users/Marcos/Desktop/Oficina/Monotributo/modelo.xlsx" # Ruta del excel de plantilla
        excel_file_path = f"C:/Users/Marcos/Desktop/Oficina/Monotributo/{(cuits.iloc[(len(cuits)-1),0]).upper()}.xlsx" # Ruta del nuevo excel a crear
        
        # Copia el archivo modelo al nuevo archivo Excel
        shutil.copyfile(source_file, excel_file_path)

        if len(alldata_emitidos) > 0:
            # Filtra y selecciona los datos de ventas del DataFrame completo.
            data_ventas = alldata_emitidos[['Fecha', 'Tipo', 'Número Desde', 'Denominación Receptor', 'Imp. Total']]

            # Convierte la columna 'Fecha' a formato de fecha.
            data_ventas['Fecha'] = pd.to_datetime(data_ventas['Fecha'], format='%d/%m/%Y')
            data_ventas['Fecha'] = data_ventas['Fecha'].dt.strftime('%d/%m/%Y') 

            # Ordena el dataframe por número de factura
            data_ventas = data_ventas.sort_values(by='Número Desde', ascending=True).reset_index(drop=True)

            # Aplica la función 'nota_de_credito' a los datos de compras.
            data_ventas = nota_de_credito(data_ventas)

            # Lee los datos existentes del archivo Excel del cliente.
            data_excel = pd.read_excel(excel_file_path, sheet_name="VENTAS NUEVO", usecols="A:E", skiprows=5)

            # Calcula el tamaño actual de los datos en el archivo Excel.
            size = len(data_excel)

            # Abre el archivo Excel y selecciona la hoja de trabajo 'VENTAS NUEVO'.
            wb = openpyxl.load_workbook(excel_file_path)
            sheet = wb['VENTAS NUEVO']

            # Agrega los datos de ventas al nuevo archivo Excel
            for row_idx, row in enumerate(openpyxl.utils.dataframe.dataframe_to_rows(data_ventas, index=False, header=False), (size+7)):
                for col_idx, value in enumerate(row, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            # Aplica formato a las celdas para mejorar la apariencia.
            for row in wb["VENTAS NUEVO"].iter_rows(min_row=(size+7), max_row=(size+len(data_ventas)+7), min_col=1, max_col=5):
                for cell in row:
                    cell.font = Font(name="Calibri", size=11) 
                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal="right")  
                    elif cell.column == 3:
                        cell.alignment = Alignment(horizontal="center")  
            wb.save(excel_file_path)
            
        if len(alldata_recibidos) > 0:
            # Filtra y selecciona los datos de compras del DataFrame completo.
            data_compras = alldata_recibidos[['Fecha', 'Tipo', 'Número Desde', 'Denominación Emisor', 'Imp. Total']]
            
            # Convierte la columna 'Fecha' a formato de fecha y ordena el dataframe.
            data_compras['Fecha'] = pd.to_datetime(data_compras['Fecha'], format='%d/%m/%Y')
            data_compras = data_compras.sort_values(by='Fecha', ascending=True).reset_index(drop=True)
            data_compras['Fecha'] = data_compras['Fecha'].dt.strftime('%d/%m/%Y') 

            # Aplica la función 'nota_de_credito' a los datos de compras.
            data_compras = nota_de_credito(data_compras)

            # Lee los datos existentes del archivo Excel del cliente.
            data_excel = pd.read_excel(excel_file_path, sheet_name="COMPRAS NUEVO", usecols="A:E", skiprows=5)

            # Calcula el tamaño actual de los datos en el archivo Excel.
            size = len(data_excel)

            # Abre el archivo Excel y selecciona la hoja de trabajo 'COMPRAS NUEVO'.
            wb = openpyxl.load_workbook(excel_file_path)
            sheet = wb['COMPRAS NUEVO']

            # Agrega los datos de compras al nuevo archivo Excel.
            for row_idx, row in enumerate(openpyxl.utils.dataframe.dataframe_to_rows(data_compras, index=False, header=False), (size+7)):
                for col_idx, value in enumerate(row, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            # Aplica formato a las celdas para mejorar la apariencia.
            for row in wb["COMPRAS NUEVO"].iter_rows(min_row=(size+7), max_row=(size+len(data_compras)+7), min_col=1, max_col=5):
                for cell in row:
                    cell.font = Font(name="Calibri", size=11) 
                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal="right")  
            wb.save(excel_file_path)


if __name__ == "__main__":
    main_menu()